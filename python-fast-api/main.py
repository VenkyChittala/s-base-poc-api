from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import FileResponse
import os
import shutil
import random
import string
import openpyxl
import uvicorn
from fastapi.middleware.cors import CORSMiddleware


app = FastAPI()
UPLOAD_FOLDER = 'uploads'
TEMP_FOLDER = 'temp'
# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins, you can restrict it to specific origins if needed
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],  # Add other HTTP methods if needed
    allow_headers=["*"],  # Allow all headers
)

def generate_unique_filename(filename):
    random_string = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
    return random_string + '_' + filename

def modify_and_export_data(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        # Modify data as needed
        # For example, modify the first cell of the first row
        sheet.cell(row=1, column=1).value = "Modified: " + str(sheet.cell(row=1, column=1).value)
        modified_file_path = os.path.join(TEMP_FOLDER, 'modified_file.xlsx')
        wb.save(modified_file_path)
        return modified_file_path
    except Exception as e:
        print('Error:', e)
        return None

@app.post("/upload/")
async def upload_file(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        filename = generate_unique_filename(file.filename)
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        with open(file_path, "wb") as f:
            f.write(contents)
        modified_file_path = modify_and_export_data(file_path)
        if modified_file_path:
            return FileResponse(path=modified_file_path, filename='modified_file.xlsx')
        else:
            raise HTTPException(status_code=500, detail="Error processing file")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == '__main__':
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    if not os.path.exists(TEMP_FOLDER):
        os.makedirs(TEMP_FOLDER)
    uvicorn.run("main:app",host="0.0.0.0",port=3000,log_level="info",reload=True)
